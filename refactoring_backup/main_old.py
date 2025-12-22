#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
LogStorm - Attendance Log Analyzer
Анализ логов посещаемости с генерацией Excel отчетов
"""

import pandas as pd
import json
import os
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Импорт конфигурации
from config import (
    LOGS_FILE, PERSON_PREFS_FILE, OUTPUT_EXCEL_FILE, OUTPUT_AI_SUMMARY_FILE,
    DEFAULT_WORK_HOURS, DEFAULT_START_TIME, DEFAULT_END_TIME,
    OVERTIME_THRESHOLD, LATE_THRESHOLD_MINUTES, CRITICAL_LATE_MINUTES,
    CRITICAL_UNDERWORK_HOURS, CRITICAL_UNDERWORK_DIFF,
    NIGHT_HOUR_START, NIGHT_HOUR_END, MASS_ABSENCE_THRESHOLD,
    CRITICAL_ABSENCE_THRESHOLD,
    HEADER_COLOR, SUMMARY_HEADER_COLOR, LATE_BG_COLOR, OVERTIME_BG_COLOR, 
    SUSPICIOUS_BG_COLOR, LATE_CELL_COLOR, UNDERWORK_CELL_COLOR, 
    OVERTIME_CELL_COLOR, SUSPICIOUS_CELL_COLOR, TECHNICAL_FILL_COLOR,
    SHEET_MAIN_REPORT, SHEET_SUSPICIOUS, SHEET_MONTH_PREFIX, 
    TOP_N_USERS, MAX_SUSPICIOUS_DETAILS,
    GIGACHAT_SCOPE, AI_PROMPT_SENTENCES, DAYS_RU, DAYS_ORDER, MONTHS_RU
)

# Загрузка переменных окружения из .env файла
try:
    from dotenv import load_dotenv
    load_dotenv()  # Загружает переменные из .env в os.environ
except ImportError:
    print("⚠️  python-dotenv не установлен. Используются системные переменные окружения.")

# GigaChat импорт (опционально)
try:
    from gigachat import GigaChat
    GIGACHAT_AVAILABLE = True
except ImportError:
    GIGACHAT_AVAILABLE = False
    print("⚠️  GigaChat не установлен. Для AI анализа выполните: pip install gigachat")


class AttendanceAnalyzer:
    """Класс для анализа посещаемости"""
    
    def __init__(self, logs_path=LOGS_FILE, prefs_path=PERSON_PREFS_FILE):
        self.logs_path = logs_path
        self.prefs_path = prefs_path
        self.df = None
        self.prefs = None
        self.report_data = []
        self.suspicious_cases = []
        
    def load_data(self):
        """Загрузка данных из CSV и JSON"""
        print("Загрузка данных...")
        
        # Загрузка логов
        self.df = pd.read_csv(self.logs_path)
        self.df['timestamp'] = pd.to_datetime(self.df['timestamp'])
        self.df['date'] = self.df['timestamp'].dt.date
        self.df['time'] = self.df['timestamp'].dt.time
        
        # Загрузка предпочтений
        with open(self.prefs_path, 'r', encoding='utf-8') as f:
            self.prefs = json.load(f)
        
        print(f"Загружено {len(self.df)} записей")
        print(f"Загружено {len(self.prefs)} профилей")
        
    def filter_known_users(self):
        """Фильтрация только пользователей с данными в person_prefs"""
        known_users = set(self.prefs.keys())
        initial_count = len(self.df)
        self.df = self.df[self.df['name'].isin(known_users)]
        filtered_count = len(self.df)
        print(f"Отфильтровано: {initial_count} -> {filtered_count} записей (только пользователи с профилями)")
        
    def get_work_schedule(self, user_name):
        """Получение графика работы пользователя, включая рабочие дни"""
        user_prefs = self.prefs.get(user_name, {})
        start_str = user_prefs.get('start_time', DEFAULT_START_TIME)
        start_time = datetime.strptime(start_str, '%H:%M').time()
        # Если end_time не указан, рассчитываем его как start_time + N часов
        if 'end_time' in user_prefs:
            end_str = user_prefs.get('end_time')
            end_time = datetime.strptime(end_str, '%H:%M').time()
        else:
            start_datetime = datetime.combine(datetime.today(), start_time)
            end_datetime = start_datetime + timedelta(hours=DEFAULT_WORK_HOURS)
            end_time = end_datetime.time()

        # Получаем индивидуальные рабочие дни или стандартные (Monday-Friday)
        workdays = user_prefs.get('workdays', ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'])
        return start_time, end_time, workdays
    
    def _detect_mass_absence_days(self):
        """Определение дней с массовым отсутствием работников (>90%)"""
        mass_absence_dates = set()
        
        # Определяем диапазон дат для проверки
        min_date = self.df['date'].min()
        max_date = self.df['date'].max()
        all_dates = pd.date_range(start=min_date, end=max_date, freq='D').date
        
        for date in all_dates:
            # Определяем день недели
            weekday = pd.Timestamp(date).day_name()
            
            # Подсчитываем, сколько пользователей должны работать в этот день
            users_should_work = 0
            for user_name in self.prefs.keys():
                _, _, workdays = self.get_work_schedule(user_name)
                if weekday in workdays:
                    users_should_work += 1
            
            # Если это рабочий день хотя бы для кого-то
            if users_should_work > 0:
                # Подсчитываем, сколько пользователей присутствовали
                users_present = len(self.df[self.df['date'] == date]['name'].unique())
                
                # Процент отсутствующих
                absence_rate = 1 - (users_present / users_should_work)
                
                # Если отсутствует более 90% работников
                if absence_rate > MASS_ABSENCE_THRESHOLD:
                    mass_absence_dates.add(date)
                    print(f"  ! Массовое отсутствие {date}: {users_present}/{users_should_work} ({absence_rate*100:.1f}% отсутствуют)")
        
        return mass_absence_dates

    def _detect_critical_absence_days(self):
        """Определение дней с критическим отсутствием (100% отсутствуют, включая выходные)"""
        critical_absence_dates = set()
        
        # Определяем диапазон дат для проверки
        min_date = self.df['date'].min()
        max_date = self.df['date'].max()
        all_dates = pd.date_range(start=min_date, end=max_date, freq='D').date
        
        total_users = len(self.prefs.keys())
        
        for date in all_dates:
            # Подсчитываем, сколько пользователей присутствовали
            users_present = len(self.df[self.df['date'] == date]['name'].unique())
            
            # Если никто не пришел (100% отсутствие)
            if users_present == 0:
                critical_absence_dates.add(date)
                print(f"  ! Критическое отсутствие {date}: 0/{total_users} (100% отсутствуют)")
        
        return critical_absence_dates

    
    def analyze_attendance(self):
        """Анализ посещаемости по дням с учетом индивидуальных рабочих дней"""
        print("\nАнализ посещаемости...")

        # Сначала определяем массовые и критические отсутствия по дням
        mass_absence_dates = self._detect_mass_absence_days()
        critical_absence_dates = self._detect_critical_absence_days()
        
        # Определяем диапазон дат для анализа
        min_date = self.df['date'].min()
        max_date = self.df['date'].max()
        all_dates = pd.date_range(start=min_date, end=max_date, freq='D').date
        
        # Получаем все уникальные пользователи
        all_users = list(self.prefs.keys())
        
        # Группируем существующие записи
        grouped = self.df.groupby(['name', 'date'])
        
        # Итерируем по всем комбинациям пользователь-дата
        for user_name in all_users:
            for date in all_dates:
                display_name = self.prefs.get(user_name, {}).get('display_name', user_name)
                
                # График работы и индивидуальные рабочие дни
                schedule_start, schedule_end, workdays = self.get_work_schedule(user_name)
                
                # День недели (например, 'Monday')
                weekday = pd.Timestamp(date).day_name()
                is_workday = weekday in workdays
                
                # Проверяем, является ли этот день днем массового/критического отсутствия
                is_mass_absence_day = date in mass_absence_dates
                is_critical_absence_day = date in critical_absence_dates
                
                # Проверяем, есть ли записи для этого пользователя в этот день
                if (user_name, date) in grouped.groups:
                    group = grouped.get_group((user_name, date))
                    
                    # Фильтруем ночные записи для определения рабочего времени
                    day_entries = group[
                        (group['timestamp'].dt.hour >= NIGHT_HOUR_END) &
                        (group['timestamp'].dt.hour < NIGHT_HOUR_START)
                    ]
                    
                    # Если есть дневные записи, используем их
                    if len(day_entries) > 0:
                        first_entry = day_entries['timestamp'].min()
                        last_entry = day_entries['timestamp'].max()
                        arrival_time = first_entry.time()
                        departure_time = last_entry.time()
                        
                        # Расчет рабочих часов (по дневным записям)
                        work_duration = datetime.combine(date, departure_time) - datetime.combine(date, arrival_time)
                        work_hours = work_duration.total_seconds() / 3600
                        
                        # Количество появлений (ТОЛЬКО дневные)
                        appearances = len(day_entries)
                    else:
                        # Если только ночные записи - игнорируем полностью
                        arrival_time = None
                        departure_time = None
                        work_hours = 0
                        appearances = 0
                else:
                    # Отсутствие - нет записей в логах
                    arrival_time = None
                    departure_time = None
                    work_hours = 0
                    appearances = 0
                
                # Расчет ожидаемой продолжительности рабочего дня
                expected_duration = datetime.combine(date, schedule_end) - datetime.combine(date, schedule_start)
                expected_hours = expected_duration.total_seconds() / 3600

                # Опоздание, недоработка - только для рабочих дней
                late = False
                late_minutes = 0
                underwork = False
                underwork_minutes = 0
                underwork_hours = False
                
                if is_workday:
                    # Опоздание (только если есть записи)
                    if arrival_time and arrival_time > schedule_start:
                        late = True
                        late_delta = datetime.combine(date, arrival_time) - datetime.combine(date, schedule_start)
                        late_minutes = late_delta.total_seconds() / 60

                    # Недоработка по времени ухода (только если есть записи)
                    if departure_time and departure_time < schedule_end:
                        underwork = True
                        underwork_delta = datetime.combine(date, schedule_end) - datetime.combine(date, departure_time)
                        underwork_minutes = underwork_delta.total_seconds() / 60

                    # Недоработка по часам (отработал меньше положенного)
                    underwork_hours = work_hours < expected_hours

                # Переработка (больше N часов)
                overtime = work_hours > OVERTIME_THRESHOLD

                # Классификация проблем
                suspicious = []  # Технические сбои системы (не вина сотрудника)
                employee_issues = []  # Реальные проблемы сотрудника
                
                # === ТЕХНИЧЕСКИЕ СБОИ ===
                
                # 0a. День критического отсутствия (100% отсутствуют, включая выходные)
                if is_critical_absence_day:
                    suspicious.append("День критического отсутствия (100% отсутствуют)")
                
                # 0b. День массового отсутствия (проверяем для всех в рабочий день)
                elif is_mass_absence_day and is_workday:
                    suspicious.append("День массового отсутствия (>90% отсутствуют)")
                
                # Если отсутствие в рабочий день и НЕТ массового/критического отсутствия
                if appearances == 0 and is_workday and not is_mass_absence_day and not is_critical_absence_day:
                    employee_issues.append("Отсутствие")

                # === ТЕХНИЧЕСКИЕ СБОИ (проверяем для всех дней с записями) ===
                
                # Ночная активность (аномалия времени)
                # В РАБОЧИЕ дни - это технический сбой (странное время)
                # В ВЫХОДНЫЕ дни - это нормально (человек может работать в любое время)
                if appearances > 0 and is_workday and (user_name, date) in grouped.groups:
                    group_check = grouped.get_group((user_name, date))
                    night_entries = group_check[
                        (group_check['timestamp'].dt.hour < NIGHT_HOUR_END) |
                        (group_check['timestamp'].dt.hour >= NIGHT_HOUR_START)
                    ]
                    if len(night_entries) > 0:
                        suspicious.append("Ночная активность")

                # === ДРУГИЕ ТЕХНИЧЕСКИЕ СБОИ И ПРОБЛЕМЫ (только для рабочих дней) ===
                if is_workday and appearances > 0:
                    # 1. Одно появление (камера не зафиксировала приход или уход)
                    if appearances == 1:
                        suspicious.append("Одно появление")

                    # 2. Экстремальное опоздание (>4 часов) = вероятно сбой
                    if late and late_minutes >= CRITICAL_LATE_MINUTES:
                        suspicious.append(
                            f"Экстремальное опоздание ({int(late_minutes//60)}ч "
                            f"{int(late_minutes%60)}м)"
                        )

                    # 3. Критическая недоработка = ВСЕГДА технический сбой (только в рабочие дни)
                    hours_difference = expected_hours - work_hours
                    is_critical_underwork = (
                        work_hours < CRITICAL_UNDERWORK_HOURS or
                        hours_difference >= CRITICAL_UNDERWORK_DIFF
                    )

                    if is_critical_underwork:
                        suspicious.append(
                            f"Критическая недоработка ({round(work_hours, 1)}ч)"
                        )

                    # === ПРОБЛЕМЫ СОТРУДНИКА (только в рабочие дни) ===
                    # Учитываем только если НЕТ технических сбоев
                    if not suspicious:
                        # Опоздание (15 мин - 4 часа)
                        if late and LATE_THRESHOLD_MINUTES < late_minutes < CRITICAL_LATE_MINUTES:
                            employee_issues.append(
                                f"Опоздание {int(late_minutes)} мин"
                            )
                        # Недоработка (до критического уровня)
                        hours_difference = expected_hours - work_hours
                        if (underwork_hours and
                            hours_difference < CRITICAL_UNDERWORK_DIFF and
                            work_hours >= CRITICAL_UNDERWORK_HOURS):
                            employee_issues.append(
                                f"Недоработка {round(hours_difference, 1)}ч"
                            )

                # === Переработка ===
                # В рабочий день: если больше N часов
                # В выходной день: если есть хоть одна запись
                if is_workday:
                    overtime = work_hours > OVERTIME_THRESHOLD
                else:
                    # В выходной - любая работа считается переработкой
                    overtime = appearances > 0

                record = {
                    'День недели': DAYS_RU.get(weekday, weekday),
                    'Дата': date,
                    'Имя': display_name,
                    'ID': user_name,
                    'Приход': arrival_time.strftime('%H:%M:%S') if arrival_time else '-',
                    'Уход': departure_time.strftime('%H:%M:%S') if departure_time else '-',
                    'Рабочих часов': round(work_hours, 2),
                    'Опоздание': 'Да' if late else 'Нет',
                    'Опоздание (мин)': int(late_minutes) if late else 0,
                    'Ранний уход': 'Да' if underwork else 'Нет',
                    'Недоработка (мин)': int(underwork_minutes) if underwork else 0,
                    'Недоработка часов': 'Да' if underwork_hours else 'Нет',
                    'Переработка': 'Да' if overtime else 'Нет',
                    'Появлений': appearances,
                    'Технический сбой': ', '.join(suspicious) if suspicious else 'Нет',
                    'Проблемы сотрудника': ', '.join(employee_issues) if employee_issues else 'Нет',
                    'График начало': schedule_start.strftime('%H:%M'),
                    'График конец': schedule_end.strftime('%H:%M'),
                    'Рабочий день': 'Да' if is_workday else 'Нет'
                }

                self.report_data.append(record)

                if suspicious:
                    self.suspicious_cases.append(record)

        print(f"Проанализировано {len(self.report_data)} записей")
        print(f"Найдено {len(self.suspicious_cases)} подозрительных случаев")
    
    def _create_styles(self):
        """Создание централизованных стилей для Excel"""
        styles = {
            # Заголовки
            'header_fill': PatternFill(
                start_color=HEADER_COLOR,
                end_color=HEADER_COLOR,
                fill_type="solid"
            ),
            'header_font': Font(bold=True, color="FFFFFF", size=11),
            'summary_header_fill': PatternFill(
                start_color=SUMMARY_HEADER_COLOR,
                end_color=SUMMARY_HEADER_COLOR,
                fill_type="solid"
            ),
            
            # Фоны строк (светлые)
            'late_fill': PatternFill(
                start_color=LATE_BG_COLOR,
                end_color=LATE_BG_COLOR,
                fill_type="solid"
            ),
            'overtime_fill': PatternFill(
                start_color=OVERTIME_BG_COLOR,
                end_color=OVERTIME_BG_COLOR,
                fill_type="solid"
            ),
            'suspicious_fill': PatternFill(
                start_color=SUSPICIOUS_BG_COLOR,
                end_color=SUSPICIOUS_BG_COLOR,
                fill_type="solid"
            ),
            
            # Ячейки (яркие)
            'late_cell_fill': PatternFill(
                start_color=LATE_CELL_COLOR,
                end_color=LATE_CELL_COLOR,
                fill_type="solid"
            ),
            'underwork_cell_fill': PatternFill(
                start_color=UNDERWORK_CELL_COLOR,
                end_color=UNDERWORK_CELL_COLOR,
                fill_type="solid"
            ),
            'overtime_cell_fill': PatternFill(
                start_color=OVERTIME_CELL_COLOR,
                end_color=OVERTIME_CELL_COLOR,
                fill_type="solid"
            ),
            'overtime_cell_font': Font(color="FFFFFF", bold=True),
            'suspicious_cell_fill': PatternFill(
                start_color=SUSPICIOUS_CELL_COLOR,
                end_color=SUSPICIOUS_CELL_COLOR,
                fill_type="solid"
            ),
            'suspicious_cell_font': Font(color="FFFFFF", bold=True),
            'technical_fill': PatternFill(
                start_color=TECHNICAL_FILL_COLOR,
                end_color=TECHNICAL_FILL_COLOR,
                fill_type="solid"
            ),
            
            # Границы
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        return styles
    
    def generate_excel_report(self, output_file=OUTPUT_EXCEL_FILE):
        """Генерация Excel отчета"""
        print(f"\nГенерация Excel отчета: {output_file}")
        
        # Создание DataFrame
        df_report = pd.DataFrame(self.report_data)
        
        # Сортировка по дате и имени
        df_report = df_report.sort_values(['Дата', 'Имя'])
        
        # Создание Excel файла
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Основной отчет
            df_report.to_excel(writer, sheet_name=SHEET_MAIN_REPORT, index=False)
            
            # Подозрительные случаи
            if self.suspicious_cases:
                df_suspicious = pd.DataFrame(self.suspicious_cases)
                df_suspicious.to_excel(
                    writer, 
                    sheet_name=SHEET_SUSPICIOUS, 
                    index=False
                )
            
            # Помесячные таблицы
            self._generate_monthly_sheets(writer, df_report)
        
        # Форматирование
        self._format_excel(output_file)
        
        print(f"Отчет сохранен: {output_file}")
    
    def _generate_monthly_sheets(self, writer, df_report):
        """Генерация помесячных таблиц с временем прихода/ухода"""
        # Преобразуем даты в pandas datetime для группировки
        df_report['Дата_dt'] = pd.to_datetime(df_report['Дата'])
        
        # Группируем по месяцам
        months = df_report['Дата_dt'].dt.to_period('M').unique()
        
        for month_period in sorted(months):
            # Фильтруем данные для конкретного месяца
            month_data = df_report[df_report['Дата_dt'].dt.to_period('M') == month_period]
            
            if month_data.empty:
                continue
            
            # Создаем структуру для таблицы
            # Строки: пользователи, Столбцы: дни месяца (по 2 колонки на день)
            users = sorted(month_data['Имя'].unique())
            dates = sorted(month_data['Дата_dt'].dt.date.unique())
            
            # Название листа
            month_name = MONTHS_RU[month_period.month]
            year = month_period.year
            sheet_name = f"{SHEET_MONTH_PREFIX}{month_name} {year}"
            
            # Создаем матрицу данных
            # Для каждого пользователя и каждой даты (2 колонки: приход и уход)
            matrix_data = []
            
            for user in users:
                row = {'Имя': user}
                user_data = month_data[month_data['Имя'] == user]
                
                for date in dates:
                    day_data = user_data[user_data['Дата_dt'].dt.date == date]
                    
                    if not day_data.empty:
                        arrival = day_data.iloc[0]['Приход']
                        departure = day_data.iloc[0]['Уход']
                        
                        # Форматируем время
                        arrival_str = arrival if pd.notna(arrival) else ''
                        departure_str = departure if pd.notna(departure) else ''
                        
                        # Две отдельные колонки для каждого дня
                        row[f"day_{date.day}_in"] = arrival_str
                        row[f"day_{date.day}_out"] = departure_str
                        
                        # Сохраняем информацию для форматирования
                        row[f"day_{date.day}_status"] = {
                            'late': day_data.iloc[0]['Опоздание'] == 'Да',
                            'early_leave': day_data.iloc[0]['Ранний уход'] == 'Да',
                            'underwork': day_data.iloc[0]['Недоработка часов'] == 'Да',
                            'overtime': day_data.iloc[0]['Переработка'] == 'Да',
                            'technical': day_data.iloc[0]['Технический сбой'] != 'Нет'
                        }
                    else:
                        row[f"day_{date.day}_in"] = ''
                        row[f"day_{date.day}_out"] = ''
                        row[f"day_{date.day}_status"] = None
                
                # Добавляем итоговые столбцы по пользователю
                user_month_data = month_data[month_data['Имя'] == user]
                row['Итого: Валидных дней'] = len(user_month_data[user_month_data['Технический сбой'] == 'Нет'])
                row['Итого: Неучтенных дней'] = len(user_month_data[user_month_data['Технический сбой'] != 'Нет'])
                row['Итого: Отработано часов'] = round(user_month_data[user_month_data['Технический сбой'] == 'Нет']['Рабочих часов'].sum(), 2)
                row['Итого: Опозданий'] = len(user_month_data[user_month_data['Опоздание'] == 'Да'])
                row['Итого: Ранних уходов'] = len(user_month_data[user_month_data['Ранний уход'] == 'Да'])
                row['Итого: Переработок'] = len(user_month_data[user_month_data['Переработка'] == 'Да'])
                
                matrix_data.append(row)
            
            # Создаем DataFrame для месяца
            df_month = pd.DataFrame(matrix_data)
            
            # Выбираем колонки с днями (без статусов)
            day_columns = []
            for date in dates:
                day_columns.append(f"day_{date.day}_in")
                day_columns.append(f"day_{date.day}_out")
            
            # Добавляем итоговые колонки
            summary_columns = [
                'Итого: Валидных дней',
                'Итого: Неучтенных дней', 
                'Итого: Отработано часов',
                'Итого: Опозданий',
                'Итого: Ранних уходов',
                'Итого: Переработок'
            ]
            
            columns_to_export = ['Имя'] + day_columns + summary_columns
            df_month_export = df_month[columns_to_export]
            
            # Записываем в Excel БЕЗ заголовков (напишем свои)
            df_month_export.to_excel(
                writer, 
                sheet_name=sheet_name, 
                index=False, 
                header=False, 
                startrow=0
            )
            
            # Сохраняем информацию о статусах для форматирования
            if not hasattr(self, 'monthly_status_data'):
                self.monthly_status_data = {}
            self.monthly_status_data[sheet_name] = {
                'dates': dates,
                'matrix_data': matrix_data
            }
            if not hasattr(self, 'monthly_status_data'):
                self.monthly_status_data = {}
            self.monthly_status_data[sheet_name] = {
                'dates': dates,
                'matrix_data': matrix_data
            }
    
    def _format_excel(self, filename):
        """Форматирование Excel файла"""
        # Создаем централизованные стили один раз
        styles = self._create_styles()
        
        wb = openpyxl.load_workbook(filename)
        
        # Форматирование основного отчета
        ws = wb[SHEET_MAIN_REPORT]
        self._apply_formatting(ws, styles)
        
        # Форматирование подозрительных случаев
        if SHEET_SUSPICIOUS in wb.sheetnames:
            ws_sus = wb[SHEET_SUSPICIOUS]
            self._apply_formatting(ws_sus, styles, highlight_all=True)
        
        # Форматирование помесячных листов
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith(SHEET_MONTH_PREFIX):
                self._format_monthly_sheet(wb[sheet_name], sheet_name, styles)
        
        wb.save(filename)
    
    def _apply_formatting(self, ws, styles, highlight_all=False):
        """Применение форматирования к листу"""
        
        # Форматирование заголовков
        for cell in ws[1]:
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = styles['border']
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Выделение строк
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.border = styles['border']
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Подсветка проблемных случаев
            if highlight_all:
                for cell in row:
                    cell.fill = styles['suspicious_fill']
                if row[14].value != 'Нет':  # Технический сбой
                    row[14].fill = styles['suspicious_cell_fill']
                    row[14].font = styles['suspicious_cell_font']
            else:
                # Технические сбои - красный фон для всей строки
                if row[14].value != 'Нет':
                    for cell in row:
                        cell.fill = styles['suspicious_fill']
                    row[13].fill = styles['suspicious_cell_fill']
                    row[14].fill = styles['suspicious_cell_fill']
                    row[14].font = styles['suspicious_cell_font']
                    continue
                
                # Проблемы сотрудника - желтая ячейка
                if row[15].value != 'Нет':
                    row[15].fill = styles['underwork_cell_fill']
                
                # Опоздания - фон строки светлый, ячейка прихода яркая
                if row[7].value == 'Да':
                    for cell in row:
                        cell.fill = styles['late_fill']
                    row[4].fill = styles['late_cell_fill']
                    row[7].fill = styles['late_cell_fill']
                    row[8].fill = styles['late_cell_fill']
                
                # Ранний уход - ячейка ухода желтая
                if row[9].value == 'Да':
                    if row[7].value != 'Да':
                        for cell in row:
                            cell.fill = styles['late_fill']
                    row[5].fill = styles['underwork_cell_fill']
                    row[9].fill = styles['underwork_cell_fill']
                    row[10].fill = styles['underwork_cell_fill']
                
                # Недоработка по часам
                if row[11].value == 'Да':
                    if row[7].value != 'Да' and row[9].value != 'Да':
                        for cell in row:
                            cell.fill = styles['late_fill']
                    row[6].fill = styles['underwork_cell_fill']
                    row[11].fill = styles['underwork_cell_fill']
                
                # Переработки - зеленый (позитив!)
                if row[12].value == 'Да':
                    for cell in row:
                        if cell.fill.start_color.rgb != styles['late_fill'].start_color.rgb:
                            cell.fill = styles['overtime_fill']
                    row[5].fill = styles['overtime_cell_fill']
                    row[5].font = styles['overtime_cell_font']
                    row[6].fill = styles['overtime_cell_fill']
                    row[6].font = styles['overtime_cell_font']
                    row[12].fill = styles['overtime_cell_fill']
                    row[12].font = styles['overtime_cell_font']
        
        # Закрепление первой строки
        ws.freeze_panes = 'A2'
    
    def _format_monthly_sheet(self, ws, sheet_name, styles):
        """Форматирование помесячного листа"""
        if not hasattr(self, 'monthly_status_data') or sheet_name not in self.monthly_status_data:
            return
        
        data_info = self.monthly_status_data[sheet_name]
        dates = data_info['dates']
        matrix_data = data_info['matrix_data']
        
        # Добавляем строки с числами и днями недели
        ws.insert_rows(1, 3)
        
        # Первая строка: числа месяца (с объединением двух колонок на каждое число)
        ws.cell(1, 1, "Дата")
        ws.cell(1, 1).fill = styles['header_fill']
        ws.cell(1, 1).font = styles['header_font']
        ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(1, 1).border = styles['border']
        
        col_idx = 2
        for date in dates:
            # Объединяем две ячейки для даты
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+1)
            ws.cell(1, col_idx, date.day)
            ws.cell(1, col_idx).fill = styles['header_fill']
            ws.cell(1, col_idx).font = styles['header_font']
            ws.cell(1, col_idx).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(1, col_idx).border = styles['border']
            ws.cell(1, col_idx+1).border = styles['border']
            col_idx += 2
        
        # Заголовок "Итог:" для итоговых столбцов (объединяем все 6 столбцов)
        summary_start_col = len(dates) * 2 + 2
        ws.merge_cells(start_row=1, start_column=summary_start_col, end_row=1, end_column=summary_start_col+5)
        ws.cell(1, summary_start_col, "Итог:")
        ws.cell(1, summary_start_col).fill = styles['summary_header_fill']
        ws.cell(1, summary_start_col).font = styles['header_font']
        ws.cell(1, summary_start_col).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(1, summary_start_col).border = styles['border']
        # Границы для остальных ячеек в объединении
        for i in range(1, 6):
            ws.cell(1, summary_start_col + i).border = styles['border']
        
        # Вторая строка: дни недели (с объединением двух колонок)
        ws.cell(2, 1, "День недели")
        ws.cell(2, 1).fill = styles['header_fill']
        ws.cell(2, 1).font = styles['header_font']
        ws.cell(2, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(2, 1).border = styles['border']
        
        col_idx = 2
        for date in dates:
            weekday_en = pd.Timestamp(date).day_name()
            weekday_ru = DAYS_RU.get(weekday_en, weekday_en)
            ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx+1)
            ws.cell(2, col_idx, weekday_ru)  # Полное название дня недели
            ws.cell(2, col_idx).fill = styles['header_fill']
            ws.cell(2, col_idx).font = styles['header_font']
            ws.cell(2, col_idx).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(2, col_idx).border = styles['border']
            ws.cell(2, col_idx+1).border = styles['border']
            col_idx += 2
        
        # Заголовки итоговых столбцов
        summary_headers = [
            'Валидных дней',
            'Неучтенных дней',
            'Отработано часов',
            'Опозданий',
            'Ранних уходов',
            'Переработок'
        ]
        for header in summary_headers:
            ws.merge_cells(start_row=2, start_column=col_idx, end_row=3, end_column=col_idx)
            ws.cell(2, col_idx, header)
            ws.cell(2, col_idx).fill = styles['summary_header_fill']
            ws.cell(2, col_idx).font = styles['header_font']
            ws.cell(2, col_idx).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(2, col_idx).border = styles['border']
            col_idx += 1
        
        # Третья строка: "Имя", "Приход", "Уход" для каждого дня
        ws.cell(3, 1, "Имя")
        ws.cell(3, 1).fill = styles['header_fill']
        ws.cell(3, 1).font = styles['header_font']
        ws.cell(3, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(3, 1).border = styles['border']
        
        col_idx = 2
        for date in dates:
            ws.cell(3, col_idx, "Приход")
            ws.cell(3, col_idx).fill = styles['header_fill']
            ws.cell(3, col_idx).font = styles['header_font']
            ws.cell(3, col_idx).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(3, col_idx).border = styles['border']
            
            ws.cell(3, col_idx+1, "Уход")
            ws.cell(3, col_idx+1).fill = styles['header_fill']
            ws.cell(3, col_idx+1).font = styles['header_font']
            ws.cell(3, col_idx+1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(3, col_idx+1).border = styles['border']
            col_idx += 2
        
        # Форматирование данных (начиная с 4 строки)
        for row_idx, user_data in enumerate(matrix_data, start=4):
            # Колонка с именем
            ws.cell(row_idx, 1).border = styles['border']
            ws.cell(row_idx, 1).alignment = Alignment(
                horizontal='left', vertical='center'
            )
            
            # Колонки с данными (приход и уход для каждого дня)
            col_idx = 2
            for date in dates:
                day_key = f"day_{date.day}"
                status_key = f"{day_key}_status"
                
                cell_in = ws.cell(row_idx, col_idx)  # Приход
                cell_out = ws.cell(row_idx, col_idx+1)  # Уход
                
                cell_in.border = styles['border']
                cell_in.alignment = Alignment(
                    horizontal='center', vertical='center'
                )
                cell_out.border = styles['border']
                cell_out.alignment = Alignment(
                    horizontal='center', vertical='center'
                )
                
                # Применяем форматирование на основе статуса
                if status_key in user_data and user_data[status_key]:
                    status = user_data[status_key]
                    
                    if status['technical']:
                        # Технический сбой - красный для обеих ячеек
                        cell_in.fill = styles['technical_fill']
                        cell_out.fill = styles['technical_fill']
                    elif status['underwork']:
                        # Недоработка - желтый для обеих ячеек
                        cell_in.fill = styles['underwork_cell_fill']
                        cell_out.fill = styles['underwork_cell_fill']
                    elif status['overtime']:
                        # Переработка - зеленый для обеих ячеек
                        cell_in.fill = styles['overtime_cell_fill']
                        cell_out.fill = styles['overtime_cell_fill']
                    elif status['late'] and status['early_leave']:
                        # Опоздание + ранний уход - желтый для обеих
                        cell_in.fill = styles['underwork_cell_fill']
                        cell_out.fill = styles['underwork_cell_fill']
                    elif status['late']:
                        # Только опоздание - желтый для прихода
                        cell_in.fill = styles['underwork_cell_fill']
                    elif status['early_leave']:
                        # Только ранний уход - желтый для ухода
                        cell_out.fill = styles['underwork_cell_fill']
                
                col_idx += 2
        
        # Автоширина колонок
        ws.column_dimensions['A'].width = 25  # Колонка с именами
        # Колонки с днями
        for col_idx_day in range(2, len(dates) * 2 + 2):
            col_letter = get_column_letter(col_idx_day)
            ws.column_dimensions[col_letter].width = 10
        # Итоговые колонки
        summary_start_col = len(dates) * 2 + 2
        for col_idx_summary in range(summary_start_col, summary_start_col + 6):
            col_letter = get_column_letter(col_idx_summary)
            ws.column_dimensions[col_letter].width = 15
        
        # Закрепление первых 3 строк и первой колонки
        ws.freeze_panes = 'B4'
    
    def generate_summary(self):
        """Генерация текстовой сводки"""
        print("\n" + "="*80)
        print("СВОДКА ПО АНАЛИЗУ ПОСЕЩАЕМОСТИ")
        print("="*80)
        
        df_report = pd.DataFrame(self.report_data)
        
        # Разделяем валидные записи и технические сбои
        df_valid = df_report[df_report['Технический сбой'] == 'Нет']
        df_technical = df_report[df_report['Технический сбой'] != 'Нет']
        
        # Общая статистика
        total_records = len(df_report)
        valid_records = len(df_valid)
        technical_issues = len(df_technical)
        unique_users = df_report['Имя'].nunique()
        unique_dates = df_report['Дата'].nunique()
        
        print(f"\n[ОБЩАЯ СТАТИСТИКА]")
        print(f"  • Всего записей: {total_records}")
        print(f"  • Валидных записей: {valid_records} ({valid_records/total_records*100:.1f}%)")
        print(f"  • Технических сбоев системы: {technical_issues}")
        print(f"  • Уникальных пользователей: {unique_users}")
        print(f"  • Дней в отчете: {unique_dates}")
        
        # Статистика по опозданиям (только валидные)
        late_count = len(df_valid[df_valid['Опоздание'] == 'Да'])
        late_percent = (late_count / valid_records * 100) if valid_records > 0 else 0
        
        print(f"\n[РЕАЛЬНЫЕ ОПОЗДАНИЯ СОТРУДНИКОВ]")
        print(f"  • Всего случаев: {late_count} ({late_percent:.1f}%)")
        
        if late_count > 0:
            avg_late = df_valid[df_valid['Опоздание'] == 'Да']['Опоздание (мин)'].mean()
            max_late = df_valid[df_valid['Опоздание'] == 'Да']['Опоздание (мин)'].max()
            print(f"  • Среднее опоздание: {avg_late:.0f} минут")
            print(f"  • Максимальное опоздание: {max_late:.0f} минут")
            
            # Топ опаздывающих
            late_users = (df_valid[df_valid['Опоздание'] == 'Да']
                         .groupby('Имя').size()
                         .sort_values(ascending=False)
                         .head(TOP_N_USERS))
            print(f"  • Топ-{TOP_N_USERS} опаздывающих:")
            for name, count in late_users.items():
                print(f"    - {name}: {count} раз(а)")
        
        # Статистика по переработкам (только валидные)
        overtime_count = len(df_valid[df_valid['Переработка'] == 'Да'])
        overtime_percent = ((overtime_count / valid_records * 100) 
                           if valid_records > 0 else 0)
        
        print(f"\n[ПЕРЕРАБОТКИ (>{OVERTIME_THRESHOLD} часов)]:")
        print(f"  • Всего случаев: {overtime_count} ({overtime_percent:.1f}%)")
        
        if overtime_count > 0:
            overtime_users = (df_valid[df_valid['Переработка'] == 'Да']
                             .groupby('Имя').size()
                             .sort_values(ascending=False)
                             .head(TOP_N_USERS))
            print(f"  • Топ-{TOP_N_USERS} по переработкам:")
            for name, count in overtime_users.items():
                print(f"    - {name}: {count} раз(а)")
        
        # Средняя продолжительность рабочего дня (только валидные)
        if valid_records > 0:
            avg_hours = df_valid['Рабочих часов'].mean()
            print(f"\n[РАБОЧИЕ ЧАСЫ (валидные записи)]:")
            print(f"  • Средняя продолжительность: {avg_hours:.2f} часов")
            print(f"  • Минимум: {df_valid['Рабочих часов'].min():.2f} часов")
            print(f"  • Максимум: {df_valid['Рабочих часов'].max():.2f} часов")
        
        # Проблемы сотрудников
        employee_issues = df_report[df_report['Проблемы сотрудника'] != 'Нет']
        print(f"\n[ПРОБЛЕМЫ СОТРУДНИКОВ]:")
        print(f"  • Всего случаев с проблемами: {len(employee_issues)}")
        
        # Технические сбои
        print(f"\n[ТЕХНИЧЕСКИЕ СБОИ СИСТЕМЫ]:")
        print(f"  • Всего подозрительных случаев: {technical_issues}")
        
        if self.suspicious_cases:
            sus_df = pd.DataFrame(self.suspicious_cases)
            print(f"\n  Детали:")
            for idx, row in sus_df.head(MAX_SUSPICIOUS_DETAILS).iterrows():
                print(f"    - {row['Дата']} | {row['Имя']}: "
                      f"{row['Технический сбой']}")
            
            if len(self.suspicious_cases) > MAX_SUSPICIOUS_DETAILS:
                remaining = len(self.suspicious_cases) - MAX_SUSPICIOUS_DETAILS
                print(f"    ... и еще {remaining} случаев")
        
        # Анализ по дням недели
        df_report['День недели'] = pd.to_datetime(df_report['Дата']).dt.day_name()
        weekday_stats = (df_report.groupby('День недели')['Опоздание']
                        .apply(lambda x: (x == 'Да').sum()))
        
        print(f"\n[ОПОЗДАНИЯ ПО ДНЯМ НЕДЕЛИ]:")
        
        for day in DAYS_ORDER:
            if day in weekday_stats.index:
                print(f"  • {DAYS_RU[day]}: {weekday_stats[day]} опозданий")
        
        print("\n" + "="*80)
        
    def generate_ai_summary(self):
        """Генерация AI-описания через GigaChat API"""
        print("\n[AI АНАЛИЗ]:")
        
        if not GIGACHAT_AVAILABLE:
            print("  ! GigaChat не установлен")
            print("  Для включения AI анализа выполните:")
            print("  pip install gigachat")
            return
        
        # Проверка наличия API ключа
        api_key = os.environ.get('GIGACHAT_API_KEY')
        if not api_key:
            print("  ! Не найден API ключ GigaChat")
            print("\n  📝 Настройка:")
            print("  1. Скопируйте файл .env.example в .env")
            print("  2. Откройте .env и вставьте ваш API ключ")
            print("  3. Запустите скрипт снова")
            print("\n  Или установите переменную окружения:")
            print("  export GIGACHAT_API_KEY='ваш_ключ'  # Linux/Mac")
            print("  set GIGACHAT_API_KEY=ваш_ключ     # Windows")
            print("\n  🔑 Получить ключ: https://developers.sber.ru/gigachat")
            print("  📚 Инструкция: см. файл GIGACHAT_SETUP.md")
            return
        
        try:
            df_report = pd.DataFrame(self.report_data)
            
            # Подготовка статистики для промпта
            total_records = len(df_report)
            unique_users = df_report['Имя'].nunique()
            
            # Реальные опоздания сотрудников (исключаем технические сбои)
            df_valid = df_report[df_report['Технический сбой'] == 'Нет']
            late_count = len(df_valid[df_valid['Опоздание'] == 'Да'])
            
            # Переработки (только валидные записи)
            overtime_count = len(df_valid[df_valid['Переработка'] == 'Да'])
            
            # Технические сбои
            technical_issues_count = len(df_report[df_report['Технический сбой'] != 'Нет'])
            
            # Реальные проблемы сотрудников
            employee_problems_count = len(df_report[df_report['Проблемы сотрудника'] != 'Нет'])
            
            # Средние рабочие часы (только валидные записи)
            avg_hours = df_valid['Рабочих часов'].mean()
            
            # Топ опаздывающих (только реальные опоздания)
            if late_count > 0:
                late_users = df_valid[df_valid['Опоздание'] == 'Да'].groupby('Имя').size().sort_values(ascending=False).head(3)
                late_top = ', '.join([f"{name} ({count})" for name, count in late_users.items()])
            else:
                late_top = "нет"
            
            # Топ по переработкам (только валидные записи)
            if overtime_count > 0:
                overtime_users = df_valid[df_valid['Переработка'] == 'Да'].groupby('Имя').size().sort_values(ascending=False).head(3)
                overtime_top = ', '.join([f"{name} ({count})" for name, count in overtime_users.items()])
            else:
                overtime_top = "нет"
            
            # Процент валидных записей
            valid_percent = (len(df_valid) / total_records * 100) if total_records > 0 else 0
            
            prompt = f"""Проанализируй данные посещаемости сотрудников и составь краткую профессиональную сводку.

ВАЖНО: Статистика очищена от технических сбоев системы. Учитываются только реальные проблемы сотрудников.

Статистика:
- Всего записей: {total_records}
- Валидных записей: {len(df_valid)} ({valid_percent:.1f}%)
- Технических сбоев системы: {technical_issues_count} (не учитываются в статистике сотрудников)
- Уникальных сотрудников: {unique_users}
- Реальных опозданий: {late_count} ({late_count/len(df_valid)*100:.1f}% от валидных)
- Топ опаздывающих: {late_top}
- Переработок: {overtime_count} ({overtime_count/len(df_valid)*100:.1f}% от валидных)
- Топ по переработкам: {overtime_top}
- Реальных проблем сотрудников: {employee_problems_count}
- Средняя продолжительность рабочего дня: {avg_hours:.2f} часов

Составь сводку на 4-5 предложений:
1. Общая оценка дисциплины (на основе валидных данных)
2. Ключевые проблемы сотрудников (если есть)
3. Позитивные моменты
4. Краткие рекомендации
5. Оценка качества работы системы учета (если много технических сбоев)

Ответь кратко и по делу, без воды."""

            print("  Отправка запроса в GigaChat...")
            
            # Инициализация GigaChat
            with GigaChat(credentials=api_key, verify_ssl_certs=False) as giga:
                response = giga.chat(prompt)
                ai_summary = response.choices[0].message.content
            
            print("\n" + "="*80)
            print(ai_summary)
            print("="*80)
            
            # Сохранение в файл
            with open(OUTPUT_AI_SUMMARY_FILE, 'w', encoding='utf-8') as f:
                f.write("AI АНАЛИЗ ПОСЕЩАЕМОСТИ\n")
                f.write("="*80 + "\n\n")
                f.write(ai_summary)
                f.write("\n\n" + "="*80 + "\n")
                f.write(f"Дата генерации: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            
            print(f"\n[OK] AI анализ сохранен в файл: {OUTPUT_AI_SUMMARY_FILE}")
            
        except Exception as e:
            print(f"  [X] Ошибка при обращении к GigaChat API: {e}")
            print("  Проверьте:")
            print("  1. Правильность API ключа")
            print("  2. Подключение к интернету")
            print("  3. Актуальность библиотеки gigachat")
        
    def run(self, output_file=OUTPUT_EXCEL_FILE):
        """Запуск полного анализа"""
        self.load_data()
        self.filter_known_users()
        self.analyze_attendance()
        self.generate_excel_report(output_file)
        self.generate_summary()
        self.generate_ai_summary()


def main():
    """Главная функция"""
    print("="*80)
    print("LogStorm - Анализатор логов посещаемости")
    print("="*80)
    
    analyzer = AttendanceAnalyzer()
    analyzer.run()
    
    print("\n[OK] Анализ завершен!")


if __name__ == '__main__':
    main()
