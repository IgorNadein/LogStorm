#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Фабрика стилей для Excel
"""

from openpyxl.styles import Font, PatternFill, Border, Side
from config import (
    HEADER_COLOR,
    SUMMARY_HEADER_COLOR,
    LATE_BG_COLOR,
    OVERTIME_BG_COLOR,
    SUSPICIOUS_BG_COLOR,
    LATE_CELL_COLOR,
    UNDERWORK_CELL_COLOR,
    OVERTIME_CELL_COLOR,
    SUSPICIOUS_CELL_COLOR,
    TECHNICAL_FILL_COLOR
)


class ExcelStyleFactory:
    """Фабрика стилей для Excel отчётов"""
    
    @staticmethod
    def create_all_styles() -> dict:
        """
        Создание всех стилей для Excel
        
        Returns:
            Словарь со всеми стилями
        """
        styles = {
            # === Заголовки ===
            'header_fill': PatternFill(
                start_color=HEADER_COLOR,
                end_color=HEADER_COLOR,
                fill_type="solid"
            ),
            'header_font': Font(bold=True, color="FFFFFF", size=11),
            
            'summary_header_fill': PatternFill(
                start_color=SUMMARY_HEADER_COLOR,
                end_color=SUMMARY_HEADER_COLOR,
                fill_type="solid"
            ),
            
            # === Фоны строк (светлые) ===
            'late_fill': PatternFill(
                start_color=LATE_BG_COLOR,
                end_color=LATE_BG_COLOR,
                fill_type="solid"
            ),
            'overtime_fill': PatternFill(
                start_color=OVERTIME_BG_COLOR,
                end_color=OVERTIME_BG_COLOR,
                fill_type="solid"
            ),
            'suspicious_fill': PatternFill(
                start_color=SUSPICIOUS_BG_COLOR,
                end_color=SUSPICIOUS_BG_COLOR,
                fill_type="solid"
            ),
            
            # === Ячейки (яркие) ===
            'late_cell_fill': PatternFill(
                start_color=LATE_CELL_COLOR,
                end_color=LATE_CELL_COLOR,
                fill_type="solid"
            ),
            'underwork_cell_fill': PatternFill(
                start_color=UNDERWORK_CELL_COLOR,
                end_color=UNDERWORK_CELL_COLOR,
                fill_type="solid"
            ),
            'overtime_cell_fill': PatternFill(
                start_color=OVERTIME_CELL_COLOR,
                end_color=OVERTIME_CELL_COLOR,
                fill_type="solid"
            ),
            'overtime_cell_font': Font(color="FFFFFF", bold=True),
            
            'suspicious_cell_fill': PatternFill(
                start_color=SUSPICIOUS_CELL_COLOR,
                end_color=SUSPICIOUS_CELL_COLOR,
                fill_type="solid"
            ),
            'suspicious_cell_font': Font(color="FFFFFF", bold=True),
            
            'technical_fill': PatternFill(
                start_color=TECHNICAL_FILL_COLOR,
                end_color=TECHNICAL_FILL_COLOR,
                fill_type="solid"
            ),
            
            # === Границы ===
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        
        return styles
