#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Форматирование Excel отчётов
"""

import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from utils import ExcelStyleFactory
from utils.smart_colors import ColorCalculator
from config import (
    SHEET_MAIN_REPORT,
    SHEET_SUSPICIOUS
)


class ExcelFormatter:
    """Форматирование Excel файлов"""
    
    def __init__(self, filename: str, reporter, color_scheme=None):
        """
        Args:
            filename: Путь к Excel файлу
            reporter: ExcelReporter с данными для форматирования
            color_scheme: ColorScheme для умных цветов (опционально)
        """
        self.filename = filename
        self.reporter = reporter
        self.styles = ExcelStyleFactory.create_all_styles()
        # Умные цвета с настраиваемой схемой
        self.color_calc = ColorCalculator(color_scheme)
    
    def format_all(self):
        """Применение форматирования ко всем листам"""
        print("Форматирование Excel...")
        
        wb = openpyxl.load_workbook(self.filename)
        
        # Основной отчёт
        if SHEET_MAIN_REPORT in wb.sheetnames:
            self._format_main_sheet(wb[SHEET_MAIN_REPORT])
        
        # Подозрительные случаи
        if SHEET_SUSPICIOUS in wb.sheetnames:
            self._format_main_sheet(wb[SHEET_SUSPICIOUS], highlight_all=True)
        
        # Месячные листы (все кроме основных)
        for sheet_name in wb.sheetnames:
            if sheet_name not in [SHEET_MAIN_REPORT, SHEET_SUSPICIOUS]:
                self._format_monthly_sheet(wb[sheet_name], sheet_name)
        
        wb.save(self.filename)
        print("Форматирование завершено")
    
    def _format_main_sheet(self, ws, highlight_all=False):
        """Форматирование основного листа"""
        # Заголовки
        for cell in ws[1]:
            cell.fill = self.styles['header_fill']
            cell.font = self.styles['header_font']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.styles['border']
        
        # Автоширина
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Данные
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.border = self.styles['border']
                cell.alignment = Alignment(
                    horizontal='left',
                    vertical='center'
                )
            
            if highlight_all:
                for cell in row:
                    cell.fill = self.styles['suspicious_fill']
                if row[14].value != 'Нет':
                    row[14].fill = self.styles['suspicious_cell_fill']
                    row[14].font = self.styles['suspicious_cell_font']
            else:
                # Проверяем рабочий день (колонка 18)
                is_workday = row[18].value == 'Да'
                
                # Технические сбои (всегда красный)
                if row[14].value != 'Нет':
                    for cell in row:
                        cell.fill = self.styles['suspicious_fill']
                    row[14].fill = self.styles['suspicious_cell_fill']
                    row[14].font = self.styles['suspicious_cell_font']
                    continue
                
                # Опоздания (только в рабочие дни)
                if is_workday and row[7].value == 'Да':
                    for cell in row:
                        cell.fill = self.styles['late_fill']
                    row[4].fill = self.styles['late_cell_fill']
                
                # Переработки (только в рабочие дни)
                if is_workday and row[12].value == 'Да':
                    late_rgb = self.styles['late_fill'].start_color.rgb
                    for cell in row:
                        if (not cell.fill or
                                cell.fill.start_color.rgb != late_rgb):
                            cell.fill = self.styles['overtime_fill']
                    row[6].fill = self.styles['overtime_cell_fill']
                    row[6].font = self.styles['overtime_cell_font']
        
        ws.freeze_panes = 'A2'
    
    def _format_monthly_sheet(self, ws, sheet_name):
        """Форматирование транспонированного листа (даты↓, имена→)"""
        if sheet_name not in self.reporter.monthly_status_data:
            return
        
        data_info = self.reporter.monthly_status_data[sheet_name]
        users = data_info['users']
        dates = data_info['dates']
        status_matrix = data_info['status_matrix']
        summary_start_row = data_info['summary_start_row']
        
        # Вставляем 2 строки для заголовков
        ws.insert_rows(1, 2)
        
        # Строка 1: "Дата" + имена пользователей (по 2 колонки)
        ws.cell(1, 1, "Дата")
        ws.cell(1, 1).fill = self.styles['header_fill']
        ws.cell(1, 1).font = self.styles['header_font']
        ws.cell(1, 1).alignment = Alignment(
            horizontal='center', vertical='center'
        )
        ws.cell(1, 1).border = self.styles['border']
        
        col_idx = 2
        for user in users:
            # Объединяем 2 колонки для имени
            ws.merge_cells(
                start_row=1, start_column=col_idx,
                end_row=1, end_column=col_idx+1
            )
            ws.cell(1, col_idx, user)
            ws.cell(1, col_idx).fill = self.styles['header_fill']
            ws.cell(1, col_idx).font = self.styles['header_font']
            ws.cell(1, col_idx).alignment = Alignment(
                horizontal='center', vertical='center'
            )
            ws.cell(1, col_idx).border = self.styles['border']
            ws.cell(1, col_idx+1).border = self.styles['border']
            col_idx += 2
        
        # Строка 2: "День недели" + Приход/Уход для каждого
        ws.cell(2, 1, "День недели")
        ws.cell(2, 1).fill = self.styles['header_fill']
        ws.cell(2, 1).font = self.styles['header_font']
        ws.cell(2, 1).alignment = Alignment(
            horizontal='center', vertical='center'
        )
        ws.cell(2, 1).border = self.styles['border']
        
        col_idx = 2
        for user in users:
            ws.cell(2, col_idx, "Приход")
            ws.cell(2, col_idx).fill = self.styles['header_fill']
            ws.cell(2, col_idx).font = self.styles['header_font']
            ws.cell(2, col_idx).alignment = Alignment(
                horizontal='center', vertical='center'
            )
            ws.cell(2, col_idx).border = self.styles['border']
            
            ws.cell(2, col_idx+1, "Уход")
            ws.cell(2, col_idx+1).fill = self.styles['header_fill']
            ws.cell(2, col_idx+1).font = self.styles['header_font']
            ws.cell(2, col_idx+1).alignment = Alignment(
                horizontal='center', vertical='center'
            )
            ws.cell(2, col_idx+1).border = self.styles['border']
            col_idx += 2
        
        # Форматирование строк с данными (данные уже в ячейках!)
        for row_idx, date in enumerate(dates, start=3):
            # Первая колонка: только форматирование, данные уже есть
            cell_date = ws.cell(row_idx, 1)
            cell_date.border = self.styles['border']
            cell_date.alignment = Alignment(
                horizontal='left', vertical='center'
            )
            
            # Данные по пользователям
            col_idx = 2
            for user in users:
                status_key = f"{date}_{user}"
                cell_in = ws.cell(row_idx, col_idx)
                cell_out = ws.cell(row_idx, col_idx+1)
                
                cell_in.border = self.styles['border']
                cell_in.alignment = Alignment(
                    horizontal='center', vertical='center'
                )
                cell_out.border = self.styles['border']
                cell_out.alignment = Alignment(
                    horizontal='center', vertical='center'
                )
                
                # Применяем стиль гиперссылки к ячейкам с формулами
                if (cell_in.value and
                        isinstance(cell_in.value, str) and
                        cell_in.value.startswith('=HYPERLINK')):
                    cell_in.style = 'Hyperlink'

                if (cell_out.value and
                        isinstance(cell_out.value, str) and
                        cell_out.value.startswith('=HYPERLINK')):
                    cell_out.style = 'Hyperlink'
                
                # Применяем цвета по статусу
                if status_key in status_matrix:
                    status = status_matrix[status_key]
                    is_workday = status.get('is_workday', True)
                    
                    # Приоритет 1: Технические сбои (цвет ошибки из схемы)
                    if status['technical']:
                        error_color = self.color_calc.scheme.error
                        cell_in.fill = PatternFill(
                            start_color=error_color,
                            end_color=error_color,
                            fill_type="solid"
                        )
                        cell_out.fill = PatternFill(
                            start_color=error_color,
                            end_color=error_color,
                            fill_type="solid"
                        )
                    
                    # Для рабочих дней: умные цвета с градациями
                    elif is_workday:
                        # Получаем числовые значения
                        late_min = status.get('late_minutes', 0)
                        early_min = status.get('early_leave_minutes', 0)
                        under_hrs = status.get('underwork_hours', 0)
                        over_hrs = status.get('overtime_hours', 0)
                        
                        # Вычисляем умные цвета
                        arrival_color, departure_color = (
                            self.color_calc.calculate_combined_color(
                                late_min, early_min, under_hrs, over_hrs
                            )
                        )
                        
                        # Применяем цвета
                        if arrival_color != "FFFFFF":
                            cell_in.fill = PatternFill(
                                start_color=arrival_color,
                                end_color=arrival_color,
                                fill_type="solid"
                            )
                        if departure_color != "FFFFFF":
                            cell_out.fill = PatternFill(
                                start_color=departure_color,
                                end_color=departure_color,
                                fill_type="solid"
                            )
                    
                    # Для выходных дней: только переработка (зеленый)
                    else:
                        if status.get('overtime', False):
                            over_hrs = status.get('overtime_hours', 0)
                            color = self.color_calc.calculate_overtime_color(
                                over_hrs
                            )
                            cell_in.fill = PatternFill(
                                start_color=color,
                                end_color=color,
                                fill_type="solid"
                            )
                            cell_out.fill = PatternFill(
                                start_color=color,
                                end_color=color,
                                fill_type="solid"
                            )
                
                col_idx += 2
        
        # Форматирование итоговых строк
        summary_start = summary_start_row + 3
        summary_count = 7  # Учтено, Неучтено, Часы, Опоздания, Уходы, Перераб, Отсутствия
        for row_idx in range(summary_start, summary_start + summary_count):
            # Первая колонка - название итога (жирный)
            ws.cell(row_idx, 1).font = self.styles['header_font']
            ws.cell(row_idx, 1).fill = self.styles['summary_header_fill']
            ws.cell(row_idx, 1).border = self.styles['border']
            ws.cell(row_idx, 1).alignment = Alignment(
                horizontal='left', vertical='center'
            )
            
            # Остальные ячейки - итоги
            for col_idx in range(2, len(users) * 2 + 2):
                cell = ws.cell(row_idx, col_idx)
                cell.border = self.styles['border']
                cell.alignment = Alignment(
                    horizontal='center', vertical='center'
                )
                cell.fill = self.styles['summary_header_fill']
                cell.font = self.styles['header_font']
        
        # Ширина колонок
        ws.column_dimensions['A'].width = 20
        for col_idx in range(2, len(users) * 2 + 2):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 10
        
        # Закрепление первой строки
        ws.freeze_panes = 'A3'
        
        ws.freeze_panes = 'B4'
